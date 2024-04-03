# Summary
# backend.py manipulates the data of an excel file and saves it to another excel file

import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

class Excel:
    def __init__(self, accountNumber, previousFile, updatedFile):
        self._previousFile = previousFile   
        self._updatedFile = updatedFile     
        self._accountNumber = accountNumber 
        self.sageCheckList = load_workbook(previousFile)
        self.keyBankImport = Workbook()
        self.preSheet = self.sageCheckList.active
        self.postSheet = self.keyBankImport.active
        self.maxRow = self.preSheet.max_row
        self.maxCol = self.preSheet.max_column
        self.checkTotal = 0
    
    @property
    def previousFile(self):
        pass

    @previousFile.setter
    def previousFile(self, newFile):
            if newFile.endswith(".xlsx"):
                self._previousFile = newFile
            else:
                print("The file needs to be a xlsx file type.")

    @property
    def accountNumber(self):
        pass

    @accountNumber.setter
    def accountNumber(self, newAccountNumber):
            self._accountNumber = newAccountNumber

    # Function applys static data into the new excel file
    def applySettings(self):
        self.postSheet['D1'] = "Check Total"
        self.postSheet['D'+str(2)] = self.checkTotal
        self.postSheet['D1'].font = openpyxl.styles.Font(bold=True, color="FF0000")
        self.postSheet['D1'].alignment = Alignment(horizontal='center', vertical='center')
        self.postSheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
        self.postSheet.column_dimensions['A'].width = 100

    # Extracts data from all the rows in the prior excel file
    def extractData(self):
        for rows in range(1, self.maxRow+1, 1):
            line = [[] for i in range(6)]
            line[0].append("00") # Region Code
            line[1].append(self._accountNumber)
            finalOutPut = None
            payeeLine = None
            for cols in 'ABCD':
                cell = "{}{}".format(cols,rows)
                if cols == 'A':
                    checkNumber = self.preSheet[cell].value
                    formatCheckNumber = str(checkNumber).zfill(10)
                    line[2].append(formatCheckNumber)
                elif cols == 'B':
                    date = self.preSheet[cell].value
                    if (date, datetime):
                            formatDate = date.strftime("%Y%m%d") + ' ' * 16 # Have Jeff Check This (Seems to be wrong at 15 spacing)
                            line[4].append(formatDate)
                    else:
                        print(f"The value is not a date: {date}")
                elif cols == 'C':
                    amount = self.preSheet[cell].value
                    self.postSheet['C'+str(rows)] = float(amount)
                    self.checkTotal += float(amount)
                    formatAmount = str(amount).replace('.', '').zfill(10)
                    line[3].append(formatAmount)
                else:
                    payeeLine = self.preSheet[cell].value
                    line[5].append(payeeLine.ljust(75,' ') + ' ' * 84)
                    solution = []
                    for i in range(0, len(line), 1):
                        for n in line[i]:
                            solution.append(n)
                    finalOutPut = ''.join(solution)
            self.postSheet['A'+str(rows)] = finalOutPut
            
            # Populate B Column with the length of the Output
            self.postSheet['B'+str(rows)] = len(finalOutPut)
        self.postSheet['D'+str(2)] = self.checkTotal

    # Function exports new excel file into Downloads directory
    def saveExcel(self):
        downloadPath = os.path.join(os.path.expanduser('~'), 'Downloads')
        file_path = os.path.join(downloadPath,self._updatedFile)
        self.keyBankImport.save(file_path)
        self.sageCheckList.close()
        self.keyBankImport.close()

accountNumber = "123456789123456"

#userSheet = Excel(accountNumber, "testdata.xlsx", "downloadtest.xlsx")
#userSheet.applySettings()
#userSheet.extractData()
#userSheet.saveExcel()